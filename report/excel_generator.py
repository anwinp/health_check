from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ExcelGenerator:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = Workbook()

    def add_table(self, table_data, headers, sheet_name="Sheet"):
        # Check if the default sheet is empty, and use it; otherwise, create a new one
        if self.workbook.active.title == "Sheet" and not self.workbook.active['A1'].value:
            sheet = self.workbook.active
            sheet.title = sheet_name
        else:
            sheet = self.workbook.create_sheet(title=sheet_name)
        
        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write table data
        for row_num, row in enumerate(table_data, start=2):
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = row.get(header, "")
                cell.alignment = Alignment(horizontal="left")

        # Adjust column widths
        for col_num, header in enumerate(headers, start=1):
            column = get_column_letter(col_num)
            sheet.column_dimensions[column].width = max(len(header), 15)

    def save(self):
        self.workbook.save(filename=self.filename)

    def get_filepath(self):
        return self.filename

